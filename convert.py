#!/usr/bin/env python3
"""
Convierte escape_rooms.xlsx → data.json
Ejecutado automáticamente por GitHub Actions.
"""

import json
import sys
import glob
import openpyxl

# ── Buscar el archivo Excel en el repo ───────────────────────────────────────
xlsx_files = glob.glob("*.xlsx") + glob.glob("*.xls")
if not xlsx_files:
    print("❌ No se encontró ningún archivo .xlsx en el repositorio.")
    sys.exit(1)

EXCEL_FILE = xlsx_files[0]
print(f"📂 Leyendo: {EXCEL_FILE}")

wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

# ── Detectar hojas ────────────────────────────────────────────────────────────
sheet_names = wb.sheetnames
pend_name  = next((s for s in sheet_names if "pend" in s.lower()), sheet_names[0])
hecho_name = next((s for s in sheet_names if any(x in s.lower() for x in ["hech","done","complet"])), sheet_names[1] if len(sheet_names) > 1 else sheet_names[0])

print(f"   Hoja Pendientes : {pend_name}")
print(f"   Hoja Hechos     : {hecho_name}")

# ── Mapa de columnas (flexible) ───────────────────────────────────────────────
COL_ALIASES = {
    "nombre":      ["nombre del escape", "nombre", "name", "escape"],
    "empresa":     ["empresa", "company"],
    "ciudad":      ["ciudad", "city", "location"],
    "tematica":    ["temática", "tematica", "theme", "tema"],
    "tipo":        ["tipo", "type"],
    "duracion":    ["duración", "duracion", "duración (min)", "duration", "min"],
    "dificultad":  ["dificultad", "difficulty"],
    "rating":      ["valoración", "valoracion", "rating escapistas", "rating", "score"],
    "web":         ["web", "url", "link"],
    "valoracion":  ["valoración grupo", "valoracion grupo", "group rating", "puntuación", "puntuacion"],
    "descripcion":   ["descripción", "descripcion", "description", "descripción del escape", "descripcion del escape", "descripción del room", "resumen"],
    "max_personas":  ["max_personas", "max personas", "máximo jugadores", "maximo jugadores", "jugadores", "players", "max players", "capacidad"],
    "historia":      ["historia", "history"],
    "ambientacion":  ["ambientación", "ambientacion", "atmosphere"],
    "jugabilidad":   ["jugabilidad", "gameplay"],
    "gamemaster":    ["gamemaster", "game master", "gm", "game_master"],
    "terpeca":       ["posición terpeca", "posicion terpeca", "terpeca", "pos terpeca", "posición_terpeca"],
}

def find_col(headers, key):
    aliases = COL_ALIASES.get(key, [key])
    for alias in aliases:
        for i, h in enumerate(headers):
            if h and str(h).strip().lower() == alias.lower():
                return i
    return -1

def safe_str(val):
    """Convierte un valor de celda a string limpio, ignorando fechas y valores no textuales."""
    if val is None:
        return ""
    import datetime
    if isinstance(val, (datetime.date, datetime.datetime)):
        return ""
    s = str(val).strip()
    import re
    if re.match(r'^\d{4}-\d{2}-\d{2}', s):
        return ""
    return s

# Campos numéricos: float para notas, int para conteos/duraciones
NUMERIC_FIELDS  = {"rating", "duracion", "valoracion", "historia", "ambientacion", "jugabilidad", "gamemaster", "terpeca"}
INTEGER_FIELDS  = {"duracion", "max_personas", "terpeca"}

def safe_num(val):
    """Convierte un valor a float limpio.
    Caso especial: Excel interpreta decimales como fechas (8.9 → 8 sep → datetime(2026,9,8))
    Se recupera el valor original como float(day.month).
    """
    if val is None:
        return ""
    import datetime
    if isinstance(val, (datetime.date, datetime.datetime)):
        # Reconstruir el decimal: día=8, mes=9 → 8.9
        return float(f"{val.day}.{val.month}")
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip().replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return ""

def parse_sheet(ws, debug_fields=None):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    col_map = {key: find_col(headers, key) for key in COL_ALIASES}

    # Debug: mostrar qué columnas se detectaron
    print(f"   Columnas detectadas: { {k: (headers[v] if v >= 0 else 'NO ENCONTRADA') for k, v in col_map.items() if v >= 0} }")

    result = []
    for row in rows[1:]:
        nombre_idx = col_map["nombre"]
        if nombre_idx < 0 or nombre_idx >= len(row):
            continue
        val = row[nombre_idx]
        if val is None or str(val).strip() == "":
            continue
        # Ignorar si parece una fila de cabecera repetida
        if str(val).strip().lower() in ["nombre del escape", "nombre", "name", "escape"]:
            print(f"   ⚠ Cabecera repetida ignorada en fila: {list(row)[:4]}")
            continue

        obj = {}
        for key, idx in col_map.items():
            val = ""
            if idx >= 0 and idx < len(row) and row[idx] is not None:
                raw = row[idx]
                if key in NUMERIC_FIELDS or key in INTEGER_FIELDS:
                    val = safe_num(raw)
                    if key in INTEGER_FIELDS and isinstance(val, float):
                        val = int(val)
                    # Debug campos numéricos problemáticos
                    if debug_fields and key in debug_fields and raw != "" and raw is not None:
                        print(f"   [{key}] raw={repr(raw)} ({type(raw).__name__}) → {repr(val)}")
                else:
                    val = safe_str(raw)
            obj[key] = val
        result.append(obj)
    return result

# ── Parsear hojas ─────────────────────────────────────────────────────────────
print(f"\n📋 Hoja Pendientes:")
pendientes_raw = parse_sheet(wb[pend_name], debug_fields={"rating"})
# Deduplicar por nombre (conservar primera aparición)
seen = set()
pendientes = []
for p in pendientes_raw:
    key = str(p.get("nombre","")).strip().lower()
    if key and key not in seen:
        seen.add(key)
        pendientes.append(p)
if len(pendientes_raw) != len(pendientes):
    print(f"   ⚠ Duplicados eliminados: {len(pendientes_raw) - len(pendientes)}")

print(f"\n📋 Hoja Hechos:")
hechos_raw2 = parse_sheet(wb[hecho_name], debug_fields={"valoracion"})
seen2 = set()
hechos_dedup = []
for h in hechos_raw2:
    key = str(h.get("nombre","")).strip().lower()
    if key and key not in seen2:
        seen2.add(key)
        hechos_dedup.append(h)
hechos = [h for h in hechos_dedup if h["nombre"]]

# Calcular ranking por valoración grupo (robusto ante valores no numéricos)
def safe_float(val):
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0

hechos_sorted = sorted(hechos, key=lambda h: safe_float(h["valoracion"]), reverse=True)
for h in hechos:
    h["ranking"] = next((i + 1 for i, x in enumerate(hechos_sorted) if x["nombre"] == h["nombre"]), 0)

# ── Escribir data.json ────────────────────────────────────────────────────────
data = {
    "pendientes": pendientes,
    "hechos": hechos,
    "meta": {
        "source": EXCEL_FILE,
        "pendientes_count": len(pendientes),
        "hechos_count": len(hechos),
    }
}

with open("data.json", "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print(f"✅ data.json generado → {len(pendientes)} pendientes, {len(hechos)} hechos.")
